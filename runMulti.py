# -*- coding: utf-8 -*-
"""
Created on Sun Oct 06 07:28:57 2019

@author: USER
"""

import numpy as np
import skfuzzy as fuzzy
import matplotlib.pyplot as plot
import mysql.connector as ms
import pandas as pd
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import islice
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.utils.dataframe import dataframe_to_rows

def getKategori():    
    conn = ms.connect(user='root', password='bevy2019', host='ec2-54-83-165-147.compute-1.amazonaws.com', database='fuzzymamdani')
    # conn = ms.connect(user='root', password='', host='localhost', database='fuzzymamdani')
    cursor = conn.cursor()
    querySelect = """SELECT * FROM variables WHERE deleted = 0"""   
    cursor.execute(querySelect)
    dataVariabel = []
    dataVariabel.append([])
    dataVariabel[0].append([])
    dataVariabel[0].append([])
    dataVariabel[0].append([])
    dataVariabel.append([])
    dataVariabel[1].append([])
    dataVariabel[1].append([])
    dataVariabel[1].append([])
    
    dataKategori = []
    dataKategori.append([]);
    dataKategori[0].append([]);
    dataKategori[0].append([]);
    
    for row in cursor:
        if(row[4] == 0): #Jenis
            dataVariabel[0][0].append(row[1]) # Nama Variabel
            dataVariabel[0][1].append([row[3], row[2]]) # Range
            dataVariabel[0][2].append(row[0]) # id
            
            dataKategori[0][0].append([]); # Deklarasi ukuran variabel dataKategori
            dataKategori[0][1].append([]);
        else:
            dataVariabel[1][0].append(row[1])
            dataVariabel[1][1].append([row[3], row[2]])
            dataVariabel[1][2].append(row[0])
    
        
    querySelect = """SELECT C.*, V.jenis FROM categories C JOIN variables V ON C.idVariabel = V.id WHERE V.deleted = 0 and C.deleted = 0"""   
    cursor.execute(querySelect)
    
    dataKategori.append([]);
    dataKategori[1].append([]);
    dataKategori[1].append([]);
    for row in cursor:
        if(row[7] == 0):
            for i in range(len(dataVariabel[0][2])):
                if(dataVariabel[0][2][i] == row[1]):
                    dataKategori[0][0][i].append(row[2])
                    dataKategori[0][1][i].append([row[3], row[4], row[5]])
                    break;
        else:
            dataKategori[1][0].append(row[2])
            dataKategori[1][1].append([row[3], row[4], row[5]])
            
    return dataVariabel, dataKategori
    conn.close()
    
def getRules():    
    conn = ms.connect(user='root', password='bevy2019', host='ec2-54-83-165-147.compute-1.amazonaws.com', database='fuzzymamdani')
    # conn = ms.connect(user='root', password='', host='localhost', database='fuzzymamdani')
    cursor = conn.cursor()
    querySelect = """SELECT rule, kategoriHasil, weight, operasi FROM rules WHERE deleted = 0"""   
    cursor.execute(querySelect)
    dataRules = []
    
    for row in cursor:
        tempRule = []
        for i in range(len(row[0].split(';')) - 1):
            if(i != len(row[0].split(';')) - 1):
                tempRule.append(int((row[0].split(';'))[i]))
        dataRules.append([tempRule, row[1], int(row[2]), row[3]])
        
    
            
    return dataRules
    conn.close()

def main(value_var_i):
    #value_var_i = [val1, val2, val3]
    #value_var_i = [8, 1, 18]
    variabel, kategori = getKategori()
    rules = getRules()
    var_independen = variabel[0][0]
    range_var_i = variabel[0][1]
    var_dependen = variabel[1][0]
    range_var_d = variabel[1][1][0]

    kategori_var_i = kategori[0][0]
    arg_kategori_i = kategori[0][1]
    kategori_var_d = kategori[1][0]
    arg_kategori_d = kategori[1][1]


    # =============================================================================
    # var_independen = ["Vonis", "Pendidikan", "Umur"];
    # range_var_i = [[0,111],[0,3],[18,111]]
    # var_dependen = ["Paket"]
    # range_var_d = [0,3]
    # 
    # kategori_var_i = []
    # arg_kategori_i = []
    # kategori_var_i.append(["Ringan","Sedang","Berat"]) #Kategori Vonis
    # arg_kategori_i.append([[0, 0, 10],[8, 11, 30],[20, 110, 110]])
    # 
    # kategori_var_i.append(["SMP","SMA","S1"]) #Kategori Pendidikan
    # arg_kategori_i.append([[0, 0, 0],[1, 1, 1],[2, 2, 2]])
    # 
    # kategori_var_i.append(["Muda","Sedang","Tua"]) #Kategori Umur
    # arg_kategori_i.append([[18, 18, 30],[30, 37, 45],[45, 110, 110]])
    # 
    # range_var_d = [0,3];
    # kategori_var_d = ["Paket1","Paket2","Paket3"]
    # arg_kategori_d = [[0, 0, 0],[1, 1, 1],[2, 2, 2]]
    # =============================================================================

    

    x_var_i = [] # x Variabel Independen
    x_var_d = [] # x Variabel Dependen
    fuzzy_f_i = []
    fuzzy_f_d = []

    # =============================================================================
    # rules = [] # Format = [[Vonis, Pendidikan, Umur], Paket, Weight, OPERASI ]
    # # =============================================================================
    # rules.append([[1, 1, 1], 1, 1, 1])
    # rules.append([[1, 1, 2], 1, 1, 1])
    # rules.append([[1, 1, 3], 2, 1, 1])
    # rules.append([[1, 2, 1], 1, 1, 1])
    # rules.append([[1, 2, 2], 1, 1, 1])
    # rules.append([[1, 2, 3], 2, 1, 1])
    # rules.append([[1, 3, 1], 1, 1, 1])
    # rules.append([[2, 1, 1], 1, 1, 1])
    # rules.append([[2, 1, 2], 2, 1, 1])
    # rules.append([[2, 1, 3], 2, 1, 1])
    # rules.append([[2, 1, 1], 2, 1, 1])
    # rules.append([[2, 2, 2], 2, 1, 1])
    # rules.append([[2, 2, 3], 2, 1, 1])
    # rules.append([[2, 2, 1], 2, 1, 1])
    # rules.append([[2, 3, 2], 3, 1, 1])
    # rules.append([[2, 3, 3], 3, 1, 1])
    # rules.append([[2, 3, 1], 2, 1, 1])
    # rules.append([[3, 1, 1], 3, 1, 1])
    # rules.append([[3, 1, 2], 3, 1, 1])
    # rules.append([[3, 1, 2], 3, 1, 1])# Dua Kali
    # rules.append([[3, 1, 3], 3, 1, 1])
    # rules.append([[3, 1, 1], 3, 1, 1])
    # rules.append([[3, 2, 2], 3, 1, 1])
    # rules.append([[3, 2, 3], 3, 1, 1])
    # rules.append([[3, 2, 1], 3, 1, 1])
    # rules.append([[3, 3, 2], 3, 1, 1])
    # rules.append([[3, 3, 3], 3, 1, 1])
    # # =============================================================================
    # =============================================================================


    for i in range(len(range_var_i)):
        x_var_i.append(np.arange(range_var_i[i][0], range_var_i[i][1], 1));
        
    x_var_d = np.arange(range_var_d[0], range_var_d[1], 1);
        
    for i in range(len(kategori_var_i)):
        temp_list = []
        for j in range(len(kategori_var_i[i])):
            temp_list.append(fuzzy.trimf(x_var_i[i], [arg_kategori_i[i][j][0], arg_kategori_i[i][j][1], arg_kategori_i[i][j][2]]))
        fuzzy_f_i.append(temp_list)
        
    for i in range(len(kategori_var_d)):
        fuzzy_f_d.append(fuzzy.trimf(x_var_d, [arg_kategori_d[i][0], arg_kategori_d[i][1], arg_kategori_d[i][2]]))
        
    member_f = []
    for i in range(len(var_independen)):
        temp_mf = []
        for j in range(len(fuzzy_f_i[i])):
            temp_mf.append(fuzzy.interp_membership(x_var_i[i], fuzzy_f_i[i][j], value_var_i[i]))
        member_f.append(temp_mf)
        
    #active_rule1 = np.amin([MF_vonis_ringan, MF_pendidikan_smp, MF_umur_muda]);        
    active_rules = []
    implikasi_rules = []
    #implikasiRule1_Paket1 = np.fmin(active_rule1, paket_paket1)
    for i in range(len(rules)):
        temp_mf = []
        for j in range(len(rules[i][0])): # Ambil Nilai Rule untuk Variabel Independen
            temp_mf.append(member_f[j][rules[i][0][j] - 1])
        active_rules.append(np.amin(temp_mf))
        implikasi_rules.append(np.fmin(np.amin(temp_mf), fuzzy_f_d[rules[i][1] - 1]))

    aggregated = implikasi_rules[0]
    for i in range(len(implikasi_rules)):
        aggregated = np.fmax(aggregated, implikasi_rules[i])
        
    idxMax = 0
    
# =============================================================================
#     paket = fuzzy.defuzz(x_var_d, aggregated, 'centroid')
#     
#     result_f = []
#     tempMax = 0
#     
#     for i in range(len(fuzzy_f_d)):
#         result_f.append(fuzzy.interp_membership(x_var_d, fuzzy_f_d[i], paket))
#         if(result_f[i] > tempMax):
#             tempMax = result_f[i]
#             idxMax = i
# =============================================================================
    #print aggregated.sum()

    if (aggregated.sum() != 0):
        paket = fuzzy.defuzz(x_var_d, aggregated, 'centroid')
    
        result_f = []
        tempMax = 0
        
        for i in range(len(fuzzy_f_d)):
            result_f.append(fuzzy.interp_membership(x_var_d, fuzzy_f_d[i], paket))
            if(result_f[i] > tempMax):
                tempMax = result_f[i]
                idxMax = i
    else: # Untuk Catch Defuzzification Error
        idxMax = 0
        

    #print("Hasil Rekomendasi = " + kategori_var_d[idxMax])
    #return kategori_var_d[idxMax]
    return idxMax

def run(input_file):
    #df = pd.read_excel(input_file)
    #print df.values
    
#    valuesExcel = df.values
#    
#    for i in range(len(valuesExcel)):
#        valuesExcel[i][5] = main([valuesExcel[i][1], valuesExcel[i][2]-1, valuesExcel[i][3], valuesExcel[i][4]-1])
#        #print valuesExcel[i][4]
#    
#    data=valuesExcel[:,1:]

#    dfResult = pd.DataFrame(data=valuesExcel[:,1:])
#    dfResult.columns = ['Vonis', 'Pendidikan', 'Umur', 'Jenis Kelamin', 'Hasil']
#    dfResult = dfResult.astype({"Hasil": int})
#    dfResult = dfResult.astype({"Hasil": str})
#    dfResult = dfResult.replace({'Hasil': {'0': 'Paket 1', '1': 'Paket 2', '2':'Paket 3', '3':'Paket 4'}})
    #dfResult.rename({0: "adsds", 1: "bdsds", 2: "cdsd", 3: "cdsd", 4: "cdsd"}, axis='columns')
    #print (dfResult)
    
    wb = load_workbook(filename = input_file)
    #print wb.active.values
    sheet = wb.active
    #print(sheet['A1'].value)
    
    dataA = sheet.values
    colsA = next(dataA)[1:]
    dataA = list(dataA)
    idxA = [r[0] for r in dataA]
    for i in range(len(dataA)):
        dataA[i] = list(dataA[i])
        dataA[i][5] = main([dataA[i][1], dataA[i][2]-1, dataA[i][3], dataA[i][4]-1])
        dataA[i] = tuple(dataA[i])
    dataA = (islice(r, 1, None) for r in dataA)
    dfA = pd.DataFrame(dataA, index=idxA, columns=colsA)
    #print dfA
    
    now = datetime.now()
    dt_string = now.strftime("%d%m%Y_%H%M%S")
    resultFilename = 'result/result_'+dt_string+'.xlsx'
    
    wbResult = Workbook(write_only=True)
    wsResult = wbResult.create_sheet()
    
    cell = WriteOnlyCell(wsResult)
    cell.style = 'Pandas'
    
    def format_first_row(row, cell):
        for c in row:
            cell.value = c
            yield cell
    
    rows = dataframe_to_rows(dfA)
    first_row = format_first_row(next(rows), cell)
    wsResult.append(first_row)
    
    for row in rows:
        row = list(row)
        cell.value = row[0]
        row[0] = cell
        wsResult.append(row)

    wbResult.save(resultFilename)
    
    #dfResult.to_excel(resultFilename, index=False)
    #dfA.to_excel(resultFilename, index=False)
    return "/"+resultFilename

#run('templateInput.xlsx')